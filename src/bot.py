"""Bot Telegram untuk mencatat lowongan magang secara otomatis."""

from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters

from src.config import EXCEL_FILE, HEADER_ROW, SHEET_NAME, TELEGRAM_TOKEN
from src.core.exceptions import (
    DataExtractionError,
    ExcelOperationError,
    ScrapingError,
    UnsupportedPlatformError,
)
from src.core.logger import logger
from src.services.extractor import extract_with_deepseek
from src.services.scraper import scrape_metadata
from src.utils.platform import get_platform_from_url


async def start(update: Update, context) -> None:
    """Handler untuk perintah /start."""
    try:
        logger.info(f"User {update.effective_user.id} started the bot")
        await update.message.reply_text(
            "Halo! Kirim link lowongan IG/TikTok/X/Threads untuk dicatat otomatis di sheet 'Data Magang'."
        )
    except Exception as e:
        logger.error(f"Error in start handler: {e}")
        await update.message.reply_text("❌ Terjadi kesalahan. Silakan coba lagi.")


async def handle_link(update: Update, context) -> None:
    """Handler untuk menerima dan memproses link lowongan."""
    try:
        url = update.message.text.strip()
        logger.info(f"Processing URL from user {update.effective_user.id}: {url}")

        supported_platforms = ["instagram.com", "tiktok.com", "x.com", "threads.net"]

        if not any(p in url for p in supported_platforms):
            logger.warning(f"Unsupported platform URL: {url}")
            raise UnsupportedPlatformError(f"Platform tidak didukung: {url}")

        await update.message.reply_text("⏳ Mengekstrak...")

        try:
            scraped = scrape_metadata(url)
            logger.info(f"Successfully scraped metadata for: {url}")
        except Exception as e:
            logger.error(f"Scraping failed for {url}: {e}")
            raise ScrapingError(f"Gagal mengambil data dari link: {e}")

        try:
            data = extract_with_deepseek(url, scraped)
            logger.info(f"Successfully extracted data for: {url}")
        except Exception as e:
            logger.error(f"Data extraction failed for {url}: {e}")
            raise DataExtractionError(f"Gagal mengekstrak informasi: {e}")

        try:
            save_to_excel(data, url)
            logger.info(f"Successfully saved data to Excel for: {url}")
        except Exception as e:
            logger.error(f"Excel save failed for {url}: {e}")
            raise ExcelOperationError(f"Gagal menyimpan ke Excel: {e}")

        reply = (
            f"✅ Tersimpan!\n"
            f"🏢 {data.get('company', '?')}\n"
            f"💼 {data.get('position', '?')}"
        )
        await update.message.reply_text(reply)
        logger.info(f"Successfully processed and saved: {url}")

    except UnsupportedPlatformError as e:
        await update.message.reply_text(
            "❌ Link tidak didukung. Gunakan Instagram, TikTok, X, atau Threads."
        )
    except (ScrapingError, DataExtractionError, ExcelOperationError) as e:
        logger.error(f"Processing error: {e}")
        await update.message.reply_text(f"❌ {e}")
    except Exception as e:
        logger.error(f"Unexpected error in handle_link: {e}")
        await update.message.reply_text(
            "❌ Terjadi kesalahan tak terduga. Silakan coba lagi."
        )


def save_to_excel(data: dict, url: str) -> None:
    """Simpan data lowongan ke file Excel.

    Args:
        data: Dictionary hasil ekstraksi informasi lowongan.
        url: URL asal lowongan.

    Raises:
        ExcelOperationError: If Excel operations fail.
    """
    logger.debug(f"Saving to {EXCEL_FILE} in sheet '{SHEET_NAME}'...")

    try:
        wb = load_workbook(EXCEL_FILE)
        logger.debug("Loaded existing Excel file")
    except FileNotFoundError:
        logger.info("Excel file not found, creating new file...")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            ws.append(HEADER_ROW)
            wb.save(EXCEL_FILE)
            wb = load_workbook(EXCEL_FILE)
            logger.info(f"Created new Excel file with sheet '{SHEET_NAME}'")
        except Exception as e:
            logger.error(f"Failed to create new Excel file: {e}")
            raise ExcelOperationError(f"Gagal membuat file Excel baru: {e}")

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        logger.debug(f"Using existing sheet '{SHEET_NAME}'")
    else:
        logger.info(f"Sheet '{SHEET_NAME}' not found, creating new sheet...")
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
        logger.info(f"Created new sheet '{SHEET_NAME}' with headers")

    # Deteksi di mana header berada
    header_row = 1
    max_scan = 5
    for row in range(1, max_scan + 1):
        first_cell = ws.cell(row=row, column=1).value
        if first_cell and str(first_cell).strip() == "No":
            header_row = row
            break
    logger.debug(f"Using header at row {header_row}")

    # Bangun indeks kolom dari baris header
    col_idx = {}
    for i, cell in enumerate(ws[header_row], start=1):
        if cell.value:
            col_name = str(cell.value).strip()
            col_idx[col_name] = i
    logger.debug(f"Found columns: {list(col_idx.keys())}")

    # Cari baris kosong pertama setelah header
    next_row = header_row + 1
    while (
        ws.cell(row=next_row, column=1).value is not None
        and str(ws.cell(row=next_row, column=1).value).strip() != ""
    ):
        next_row += 1
    logger.debug(f"Writing to row {next_row}")

    def set_cell(col_name: str, value) -> None:
        """Set nilai cell berdasarkan nama kolom."""
        if col_name in col_idx:
            # Convert lists to strings for Excel compatibility
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value)
                logger.debug(
                    f"Converted list to string for column '{col_name}': {value}"
                )
            elif value is None:
                value = ""
            ws.cell(row=next_row, column=col_idx[col_name], value=value)
        else:
            logger.warning(f"Column '{col_name}' not found in Excel sheet")

    platform = data.get("platform") or get_platform_from_url(url)

    set_cell("No", next_row - header_row)
    set_cell("Nama Perusahaan", data.get("company"))
    set_cell("Posisi / Jabatan", data.get("position"))
    set_cell("Platform Daftar", platform)
    set_cell("Tgl Daftar", data.get("deadline", ""))
    set_cell("Link Lowongan", url)
    set_cell("Catatan", data.get("description", ""))
    set_cell("Status", "Baru")
    set_cell("Divisi / Tim", data.get("division"))
    set_cell("Kota / Lokasi", data.get("location"))
    set_cell("Tipe Kerja", data.get("work_type"))
    set_cell("Kontak Rekruter", data.get("contact"))

    try:
        wb.save(EXCEL_FILE)
        logger.info("✅ Data berhasil disimpan ke Excel.")
    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")
        raise ExcelOperationError(f"Gagal menyimpan file Excel: {e}")


def main() -> None:
    """Entry point utama untuk menjalankan bot."""
    try:
        logger.info("Starting Magang Tracker Bot...")
        app = Application.builder().token(TELEGRAM_TOKEN).build()
        app.add_handler(CommandHandler("start", start))
        app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_link))
        logger.info("Bot is running and ready to receive messages...")
        app.run_polling()
    except Exception as e:
        logger.error(f"Failed to start bot: {e}")
        raise


if __name__ == "__main__":
    main()
